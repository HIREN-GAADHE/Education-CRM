"""
Student Import/Export Utilities

Handles parsing and exporting student data from/to CSV and Excel files.
"""
import pandas as pd
import io
from typing import List, Dict, Any, Tuple, Optional
from datetime import date, datetime
import logging

logger = logging.getLogger(__name__)


def parse_csv_file(file_content: bytes) -> Tuple[List[Dict[str, Any]], List[str]]:
    """
    Parse CSV file content into student dictionaries.
    
    Args:
        file_content: Raw CSV file bytes
        
    Returns:
        Tuple of (parsed_records, errors)
    """
    try:
        df = pd.read_csv(io.BytesIO(file_content))
        return _parse_dataframe(df)
    except Exception as e:
        logger.error(f"CSV parsing error: {str(e)}")
        return [], [f"CSV parsing failed: {str(e)}"]


def parse_excel_file(file_content: bytes) -> Tuple[List[Dict[str, Any]], List[str]]:
    """
    Parse Excel file content into student dictionaries.
    
    Args:
        file_content: Raw Excel file bytes
        
    Returns:
        Tuple of (parsed_records, errors)
    """
    try:
        df = pd.read_excel(io.BytesIO(file_content), engine='openpyxl')
        return _parse_dataframe(df)
    except Exception as e:
        logger.error(f"Excel parsing error: {str(e)}")
        return [], [f"Excel parsing failed: {str(e)}"]


def _parse_dataframe(df: pd.DataFrame) -> Tuple[List[Dict[str, Any]], List[str]]:
    """
    Convert pandas DataFrame to list of student dictionaries.
    
    Args:
        df: Pandas DataFrame with student data
        
    Returns:
        Tuple of (parsed_records, errors)
    """
    import numpy as np
    records = []
    errors = []
    
    # Replace NaN with None for proper JSON serialization
    df = df.where(pd.notna(df), None)
    
    # Convert column names to lowercase and strip whitespace
    df.columns = df.columns.str.strip().str.lower()
    
    # Define fields that should be strings (especially phone numbers and pincodes)
    string_fields = [
        'phone', 'alternate_phone', 'father_phone', 'mother_phone', 'guardian_phone',
        'pincode', 'roll_number', 'admission_number'
    ]
    
    for idx, row in df.iterrows():
        try:
            record = {}
            
            # Map DataFrame columns to student model fields
            for col in df.columns:
                value = row[col]
                
                # Handle None/NaN values - skip them for optional fields
                if value is None or (isinstance(value, float) and np.isnan(value)):
                    continue
                    
                # Skip empty strings
                if isinstance(value, str) and not value.strip():
                    continue
                    
                # Convert date fields
                if col in ['date_of_birth', 'admission_date']:
                    if isinstance(value, str):
                        try:
                            # Try various date formats
                            for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y']:
                                try:
                                    parsed_date = datetime.strptime(value, fmt).date()
                                    record[col] = parsed_date.isoformat()
                                    break
                                except ValueError:
                                    continue
                        except Exception:
                            errors.append(f"Row {idx + 2}: Invalid date format for {col}: {value}")
                            continue
                    elif isinstance(value, (datetime, date)):
                        record[col] = value.isoformat() if isinstance(value, date) else value.date().isoformat()
                    else:
                        record[col] = value
                        
                # Convert numeric fields to integers
                elif col in ['semester', 'year']:
                    if value is not None:
                        try:
                            record[col] = int(value)
                        except (ValueError, TypeError):
                            errors.append(f"Row {idx + 2}: Invalid numeric value for {col}: {value}")
                            continue
                
                # Convert specific numeric fields to strings (phone numbers, pincode, etc.)
                elif col in string_fields:
                    if value is not None:
                        # Convert to string, handling both int and float types
                        if isinstance(value, (int, float)):
                            # For phone numbers and numeric codes, convert to string without decimals
                            record[col] = str(int(value)) if isinstance(value, float) else str(value)
                        else:
                            record[col] = str(value).strip()
                            
                # String fields - strip whitespace and handle NaN
                elif isinstance(value, str):
                    record[col] = value.strip()
                elif isinstance(value, (int, float)):
                    # Convert other numeric values to strings if they end up in string fields
                    record[col] = str(int(value)) if isinstance(value, float) and value == int(value) else str(value)
                else:
                    record[col] = value
            
            # Only add record if it has required fields
            if record.get('admission_number') and record.get('first_name') and record.get('last_name'):
                records.append(record)
            else:
                missing = []
                if not record.get('admission_number'):
                    missing.append('admission_number')
                if not record.get('first_name'):
                    missing.append('first_name')
                if not record.get('last_name'):
                    missing.append('last_name')
                errors.append(f"Row {idx + 2}: Missing required fields: {', '.join(missing)}")
                
        except Exception as e:
            errors.append(f"Row {idx + 2}: Error processing row: {str(e)}")
            logger.error(f"Error parsing row {idx}: {str(e)}")
    
    return records, errors


def validate_student_data(record: Dict[str, Any]) -> List[str]:
    """
    Validate a single student record.
    
    Args:
        record: Student data dictionary
        
    Returns:
        List of validation errors (empty if valid)
    """
    errors = []
    
    # Required fields
    if not record.get('admission_number'):
        errors.append("admission_number is required")
    if not record.get('first_name'):
        errors.append("first_name is required")
    if not record.get('last_name'):
        errors.append("last_name is required")
    
    # Email validation (basic)
    if record.get('email') and '@' not in str(record['email']):
        errors.append("Invalid email format")
    if record.get('parent_email') and '@' not in str(record['parent_email']):
        errors.append("Invalid parent_email format")
    
    # Gender validation
    if record.get('gender') and record['gender'].lower() not in ['male', 'female', 'other']:
        errors.append("Gender must be 'male', 'female', or 'other'")
    
    return errors


def export_students_to_csv(students: List[Dict[str, Any]]) -> bytes:
    """
    Export student records to CSV format.
    
    Args:
        students: List of student dictionaries
        
    Returns:
        CSV file content as bytes
    """
    if not students:
        # Return empty CSV with headers
        df = pd.DataFrame(columns=_get_export_columns())
    else:
        df = pd.DataFrame(students)
        # Reorder columns to match template
        df = df.reindex(columns=_get_export_columns(), fill_value='')
    
    # Convert to CSV
    output = io.BytesIO()
    df.to_csv(output, index=False, encoding='utf-8')
    output.seek(0)
    return output.getvalue()


def export_students_to_excel(students: List[Dict[str, Any]]) -> bytes:
    """
    Export student records to Excel format.
    
    Args:
        students: List of student dictionaries
        
    Returns:
        Excel file content as bytes
    """
    if not students:
        # Return empty Excel with headers
        df = pd.DataFrame(columns=_get_export_columns())
    else:
        df = pd.DataFrame(students)
        # Reorder columns to match template
        df = df.reindex(columns=_get_export_columns(), fill_value='')
    
    # Convert to Excel
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Students', index=False)
    output.seek(0)
    return output.getvalue()


def create_import_template() -> bytes:
    """
    Create an Excel template file for student import with sample data.
    Includes class linking and fee assignment fields.
    
    Returns:
        Excel file content as bytes with headers and sample row
    """
    from openpyxl.utils import get_column_letter
    
    # Sample data - includes class and fee fields
    sample_data = {
        'admission_number': ['ADM001'],
        'roll_number': ['ROLL001'],
        'first_name': ['John'],
        'middle_name': [''],
        'last_name': ['Doe'],
        'date_of_birth': ['2005-01-15'],
        'gender': ['male'],
        'blood_group': ['O+'],
        'nationality': ['Indian'],
        'religion': [''],
        'caste': [''],
        'category': ['General'],
        'email': ['john.doe@example.com'],
        'phone': ['+919876543210'],
        'alternate_phone': [''],
        'address_line1': ['123 Main Street'],
        'address_line2': ['Apartment 4B'],
        'city': ['Mumbai'],
        'state': ['Maharashtra'],
        'pincode': ['400001'],
        'country': ['India'],
        'parent_email': ['parent@example.com'],
        'father_name': ['Robert Doe'],
        'father_phone': ['+919876543211'],
        'father_occupation': ['Engineer'],
        'mother_name': ['Jane Doe'],
        'mother_phone': ['+919876543212'],
        'mother_occupation': ['Teacher'],
        'guardian_name': [''],
        'guardian_phone': [''],
        'guardian_relation': [''],
        'course': ['Science'],
        'department': ['Physics'],
        'batch': ['2023-2024'],
        'class_name': ['10'],        # Class name (must match existing class in system)
        'section': ['A'],            # Section (combined with class_name to link to SchoolClass)
        'semester': [1],
        'year': [1],
        'admission_date': ['2023-06-01'],
        'admission_type': ['Regular'],
        'status': ['active'],
        'avatar_url': [''],
        # Fee assignment fields (optional)
        'fee_type': ['tuition'],     # Options: tuition, admission, examination, library, sports, transport, hostel, other
        'fee_amount': ['50000'],     # Total fee amount to assign
        'academic_year': ['2024-25'], # Academic year for the fee
        'fee_due_date': ['2024-06-30'] # Due date in YYYY-MM-DD format
    }
    
    df = pd.DataFrame(sample_data)
    
    # Create Excel file
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Student Template', index=False)
        
        # Auto-adjust column widths using proper column letter conversion
        worksheet = writer.sheets['Student Template']
        for idx, col in enumerate(df.columns):
            max_length = max(len(str(col)), 15)
            # Use get_column_letter to properly handle columns beyond Z (e.g., AA, AB, AC)
            column_letter = get_column_letter(idx + 1)  # openpyxl uses 1-based indexing
            worksheet.column_dimensions[column_letter].width = max_length
    
    output.seek(0)
    return output.getvalue()


def _get_export_columns() -> List[str]:
    """Get the standard column order for exports."""
    return [
        'admission_number', 'roll_number', 'first_name', 'middle_name', 'last_name',
        'date_of_birth', 'gender', 'blood_group', 'nationality', 'religion', 'caste', 'category',
        'email', 'phone', 'alternate_phone',
        'address_line1', 'address_line2', 'city', 'state', 'pincode', 'country',
        'parent_email', 'father_name', 'father_phone', 'father_occupation',
        'mother_name', 'mother_phone', 'mother_occupation',
        'guardian_name', 'guardian_phone', 'guardian_relation',
        'course', 'department', 'batch', 'class_name', 'section', 'semester', 'year',
        'admission_date', 'admission_type', 'status', 'avatar_url',
        'fee_type', 'fee_amount', 'academic_year', 'fee_due_date'
    ]
